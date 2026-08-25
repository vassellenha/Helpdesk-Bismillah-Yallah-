#!/usr/bin/env python3
"""Deteksi otomatis kondisi UAT: lokasi berkas, test case berikutnya, nomor screenshot."""
import os, sys, glob, subprocess

HOME = os.path.expanduser("~")

def cari(pola_list, deskripsi):
    hasil = []
    for pola in pola_list:
        hasil += glob.glob(pola, recursive=True)
    hasil = [h for h in hasil if "backup" not in os.path.basename(h).lower()
             and not os.path.basename(h).startswith("~$")]
    if not hasil:
        print(f"  [!] {deskripsi} TIDAK KETEMU. Tanyakan lokasinya ke pemilik berkas.")
        return None
    # Utamakan salinan kerja (berakhiran "Terisi"), bukan berkas kosong asli
    kerja = [h for h in hasil if "terisi" in os.path.basename(h).lower()]
    pilih = sorted(kerja or hasil)[0]
    if len(hasil) > 1:
        print(f"  [!] {deskripsi} ketemu lebih dari satu — pastikan yang dipilih benar:")
        for h in sorted(hasil):
            print(f"      {'-> ' if h == pilih else '   '}{h}")
    return pilih

print("=" * 68)
print("1. LOKASI BERKAS")
print("=" * 68)

try:
    proyek = subprocess.check_output(["git", "rev-parse", "--show-toplevel"],
                                     stderr=subprocess.DEVNULL).decode().strip()
except Exception:
    proyek = os.getcwd()
print(f"  Proyek        : {proyek}")

excel = cari([f"{HOME}/Downloads/**/Form UAT*.xlsx", f"{HOME}/Desktop/**/Form UAT*.xlsx",
              f"{HOME}/Documents/**/Form UAT*.xlsx", f"{proyek}/**/Form UAT*.xlsx"], "Form UAT")
print(f"  Form UAT      : {excel}")

folder = cari([f"{HOME}/Downloads/UAT Helpdesk*", f"{HOME}/Desktop/UAT Helpdesk*",
               f"{HOME}/Documents/UAT Helpdesk*"], "Folder screenshot")
print(f"  Folder bukti  : {folder}")

env = os.path.join(proyek, ".env")
db = "(.env tidak ada)"
if os.path.exists(env):
    for baris in open(env):
        if baris.startswith("DB_DATABASE="):
            db = baris.split("=", 1)[1].strip()
print(f"  Database      : {db}")

if excel:
    kunci = os.path.join(os.path.dirname(excel), "~$" + os.path.basename(excel))
    if os.path.exists(kunci):
        print("\n  *** EXCEL SEDANG DIBUKA. JANGAN MENULIS. Minta pemiliknya tutup dulu. ***")
        sys.exit(1)
    print("  Excel tertutup — aman untuk ditulis.")

if not excel:
    sys.exit(1)

import openpyxl
wb = openpyxl.load_workbook(excel, data_only=True)
ws = wb["Test Case UAT"]

kosong, terisi = [], []
for r in range(8, 81):
    if ws[f"B{r}"].value is None:
        continue
    status = ws[f"K{r}"].value
    (terisi if status and str(status).strip() not in ("", "Not Tested") else kosong).append(r)

print()
print("=" * 68)
print("2. PROGRES")
print("=" * 68)
print(f"  Sudah berstatus : {len(terisi)} test case")
print(f"  Belum berstatus : {len(kosong)} test case")

if not kosong:
    print("\n  SEMUA TEST CASE SUDAH TERISI.")
    sys.exit(0)

target = max(terisi) + 1 if terisi else kosong[0]
if target not in kosong:
    target = min(k for k in kosong if k > max(terisi)) if terisi else kosong[0]

bolong = [r for r in kosong if r < target]
if bolong:
    print(f"\n  [!] Baris {bolong} dilewati lebih awal — kemungkinan SENGAJA")
    print("      (mis. Login SSO belum terpasang di lokal). JANGAN diisi sendiri;")
    print("      tanyakan dulu ke pemilik berkas apakah memang dilewati.")

print()
print("=" * 68)
print(f"3. TEST CASE BERIKUTNYA — BARIS {target}")
print("=" * 68)
label = {"B": "No", "C": "Kode FR", "D": "Use Case", "E": "Aktor",
         "F": "Skenario", "G": "Langkah", "H": "Data Uji", "I": "Expected"}
for kol, nama in label.items():
    nilai = ws[f"{kol}{target}"].value
    if nilai:
        print(f"\n--- {nama} ---\n{nilai}")

uc = str(ws[f"D{target}"].value or "").split()[0]
aktor = str(ws[f"E{target}"].value or "").strip().replace("/", " : ")

print()
print("=" * 68)
print("4. PENAMAAN SCREENSHOT")
print("=" * 68)
if folder:
    tujuan = os.path.join(folder, aktor)
    print(f"  Folder tujuan : {tujuan}")
    if not os.path.isdir(tujuan):
        print(f"  (folder aktor '{aktor}' belum ada — buat dulu)")
    ada = glob.glob(os.path.join(folder, "*", f"{uc}-*.png"))
    urut = []
    for a in ada:
        ekor = os.path.basename(a).replace(f"{uc}-", "").replace(".png", "")
        if ekor.isdigit(): urut.append(int(ekor))
    mulai = max(urut) + 1 if urut else 1
    print(f"  Sudah terpakai: {uc}-01 s/d {uc}-{max(urut):02d}" if urut
          else f"  Belum ada screenshot untuk {uc}")
    print(f"  Mulai dari    : {uc}-{mulai:02d}.png")
